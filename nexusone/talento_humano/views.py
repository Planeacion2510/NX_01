from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views import View
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Count, Avg, Sum, F
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    Empleado, Contrato, Certificacion, EPS, AFP, ARL, CajaCompensacion,
    PerfilCargo, Vacante, Candidato, ProcesoSeleccion,
    Capacitacion, InscripcionCapacitacion,
    MatrizRiesgo, ExamenMedico, AccidenteTrabajo, ElementoProteccion, EntregaEPP,
    ActividadBienestar, EncuestaClimaOrganizacional, RespuestaEncuesta,
    EvaluacionDesempeño, Permiso, Vacacion, Incapacidad, Memorando, ReglamentoInterno
)

from .forms import (
    EmpleadoForm, EmpleadoBusquedaForm, ContratoForm, RenovarContratoForm, CertificacionForm,
    PerfilCargoForm, VacanteForm, CandidatoForm, ProcesoSeleccionForm,
    CapacitacionForm, InscripcionCapacitacionForm,
    MatrizRiesgoForm, ExamenMedicoForm, AccidenteTrabajoForm, ElementoProteccionForm, EntregaEPPForm,
    ActividadBienestarForm, EncuestaClimaForm, RespuestaEncuestaForm,
    EvaluacionDesempeñoForm, PermisoForm, VacacionForm, IncapacidadForm, MemorandoForm
)

# ============================================================================
# DASHBOARD Y MENÚ PRINCIPAL
# ============================================================================

@login_required
def dashboard_talento_humano(request):
    """Dashboard principal de talento humano"""
    
    # Estadísticas generales
    total_empleados = Empleado.objects.filter(estado='activo').count()
    total_inactivos = Empleado.objects.filter(estado='inactivo').count()
    
    # Contratos por vencer (próximos 30 días)
    fecha_limite = date.today() + timedelta(days=30)
    contratos_por_vencer = Contrato.objects.filter(
        activo=True,
        fecha_fin__lte=fecha_limite,
        fecha_fin__gte=date.today()
    ).count()
    
    # Vacantes abiertas
    vacantes_abiertas = Vacante.objects.filter(estado='abierta').count()
    
    # Capacitaciones programadas (próximo mes)
    capacitaciones_proximas = Capacitacion.objects.filter(
        fecha_programada__gte=date.today(),
        fecha_programada__lte=fecha_limite,
        completada=False
    ).count()
    
    # Alertas SST
    examenes_pendientes = ExamenMedico.objects.filter(
        fecha__lte=date.today() - timedelta(days=365),
        empleado__estado='activo'
    ).count()
    
    # Cumpleaños del mes
    mes_actual = date.today().month
    cumpleañeros = Empleado.objects.filter(
        fecha_nacimiento__month=mes_actual,
        estado='activo'
    ).order_by('fecha_nacimiento__day')[:5]
    
    # Empleados nuevos (últimos 30 días)
    fecha_inicio = date.today() - timedelta(days=30)
    empleados_nuevos = Empleado.objects.filter(
        fecha_ingreso__gte=fecha_inicio
    ).order_by('-fecha_ingreso')[:5]
    
    context = {
        'total_empleados': total_empleados,
        'total_inactivos': total_inactivos,
        'contratos_por_vencer': contratos_por_vencer,
        'vacantes_abiertas': vacantes_abiertas,
        'capacitaciones_proximas': capacitaciones_proximas,
        'examenes_pendientes': examenes_pendientes,
        'cumpleañeros': cumpleañeros,
        'empleados_nuevos': empleados_nuevos,
    }
    
    return render(request, 'talento_humano/dashboard.html', context)


# ============================================================================
# 1. ADMINISTRACIÓN DE PERSONAL - EMPLEADOS
# ============================================================================

class EmpleadoListView(LoginRequiredMixin, ListView):
    """Lista de empleados con búsqueda y filtros"""
    model = Empleado
    template_name = 'talento_humano/empleados/lista_empleados.html'
    context_object_name = 'empleados'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Empleado.objects.all().select_related('jefe_inmediato', 'eps', 'afp', 'arl')
        
        # Búsqueda
        busqueda = self.request.GET.get('busqueda')
        if busqueda:
            queryset = queryset.filter(
                Q(primer_nombre__icontains=busqueda) |
                Q(segundo_nombre__icontains=busqueda) |
                Q(primer_apellido__icontains=busqueda) |
                Q(segundo_apellido__icontains=busqueda) |
                Q(numero_documento__icontains=busqueda) |
                Q(cargo__icontains=busqueda)
            )
        
        # Filtros
        area = self.request.GET.get('area')
        if area:
            queryset = queryset.filter(area=area)
        
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        cargo = self.request.GET.get('cargo')
        if cargo:
            queryset = queryset.filter(cargo__icontains=cargo)
        
        return queryset.order_by('-fecha_ingreso')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_busqueda'] = EmpleadoBusquedaForm(self.request.GET)
        
        # Obtener áreas únicas para filtro
        context['areas'] = Empleado.objects.values_list('area', flat=True).distinct()
        
        # Estadísticas
        context['total_empleados'] = self.get_queryset().count()
        context['total_activos'] = self.get_queryset().filter(estado='activo').count()
        
        return context


class EmpleadoDetailView(LoginRequiredMixin, DetailView):
    """Detalle completo de empleado"""
    model = Empleado
    template_name = 'talento_humano/empleados/perfil_empleado.html'
    context_object_name = 'empleado'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empleado = self.object
        
        # Contratos
        context['contratos'] = empleado.contratos.all().order_by('-fecha_inicio')
        context['contrato_actual'] = empleado.get_contrato_actual()
        
        # Capacitaciones
        context['capacitaciones'] = InscripcionCapacitacion.objects.filter(
            empleado=empleado
        ).select_related('capacitacion').order_by('-fecha_inscripcion')[:5]
        
        # Evaluaciones de desempeño
        context['evaluaciones'] = empleado.evaluaciones.all().order_by('-fecha_evaluacion')[:3]
        
        # Permisos y vacaciones
        context['permisos_recientes'] = empleado.permisos.all().order_by('-fecha_solicitud')[:5]
        context['vacaciones_recientes'] = empleado.vacaciones.all().order_by('-fecha_solicitud')[:5]
        
        # SST
        context['examenes_medicos'] = empleado.examenes_medicos.all().order_by('-fecha')[:3]
        context['accidentes'] = empleado.accidentes.all().order_by('-fecha_accidente')[:3]
        
        # Antigüedad
        context['antiguedad'] = empleado.get_antiguedad()
        
        # Días de vacaciones disponibles
        context['dias_vacaciones_disponibles'] = empleado.get_dias_vacaciones_disponibles()
        
        return context


class EmpleadoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo empleado"""
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'talento_humano/empleados/form_empleado.html'
    success_url = reverse_lazy('talento_humano:lista_empleados')
    
    def form_valid(self, form):
        messages.success(self.request, 'Empleado creado exitosamente.')
        return super().form_valid(form)


class EmpleadoUpdateView(LoginRequiredMixin, UpdateView):
    """Editar empleado"""
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'talento_humano/empleados/form_empleado.html'
    
    def get_success_url(self):
        return reverse_lazy('talento_humano:detalle_empleado', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Empleado actualizado exitosamente.')
        return super().form_valid(form)


@login_required
def inactivar_empleado(request, pk):
    """Inactivar empleado"""
    empleado = get_object_or_404(Empleado, pk=pk)
    
    if request.method == 'POST':
        empleado.estado = 'inactivo'
        empleado.fecha_retiro = date.today()
        empleado.save()
        
        messages.success(request, f'{empleado.get_nombre_completo()} ha sido inactivado.')
        return redirect('talento_humano:lista_empleados')
    
    return render(request, 'talento_humano/empleados/confirmar_inactivacion.html', {'empleado': empleado})


# ============================================================================
# CONTRATOS
# ============================================================================

class ContratoListView(LoginRequiredMixin, ListView):
    """Lista de contratos"""
    model = Contrato
    template_name = 'talento_humano/contratos/lista_contratos.html'
    context_object_name = 'contratos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Contrato.objects.all().select_related('empleado')
        
        # Filtrar por estado
        estado = self.request.GET.get('estado')
        if estado == 'activos':
            queryset = queryset.filter(activo=True)
        elif estado == 'vencidos':
            queryset = queryset.filter(activo=False)
        
        # Filtrar por próximos a vencer
        if self.request.GET.get('por_vencer'):
            fecha_limite = date.today() + timedelta(days=30)
            queryset = queryset.filter(
                activo=True,
                fecha_fin__lte=fecha_limite,
                fecha_fin__gte=date.today()
            )
        
        return queryset.order_by('-fecha_inicio')


class ContratoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo contrato"""
    model = Contrato
    form_class = ContratoForm
    template_name = 'talento_humano/contratos/form_contrato.html'
    success_url = reverse_lazy('talento_humano:lista_contratos')
    
    def form_valid(self, form):
        messages.success(self.request, 'Contrato creado exitosamente.')
        return super().form_valid(form)


@login_required
def renovar_contrato(request, pk):
    """Renovar contrato existente"""
    contrato_anterior = get_object_or_404(Contrato, pk=pk)
    
    if request.method == 'POST':
        form = RenovarContratoForm(request.POST)
        if form.is_valid():
            # Inactivar contrato anterior
            contrato_anterior.activo = False
            contrato_anterior.save()
            
            # Crear nuevo contrato
            nuevo_contrato = Contrato.objects.create(
                empleado=contrato_anterior.empleado,
                tipo_contrato=form.cleaned_data['tipo_contrato'],
                fecha_inicio=form.cleaned_data['fecha_inicio'],
                fecha_fin=form.cleaned_data['fecha_fin'],
                salario_basico=form.cleaned_data['salario_basico'],
                auxilio_transporte=form.cleaned_data['auxilio_transporte'],
                activo=True
            )
            
            messages.success(request, f'Contrato renovado exitosamente para {contrato_anterior.empleado.get_nombre_completo()}.')
            return redirect('talento_humano:lista_contratos')
    else:
        form = RenovarContratoForm(initial={
            'tipo_contrato': contrato_anterior.tipo_contrato,
            'salario_basico': contrato_anterior.salario_basico,
            'auxilio_transporte': contrato_anterior.auxilio_transporte,
        })
    
    context = {
        'form': form,
        'contrato_anterior': contrato_anterior,
    }
    
    return render(request, 'talento_humano/contratos/renovar_contrato.html', context)


# ============================================================================
# CERTIFICACIONES
# ============================================================================

@login_required
def generar_certificacion(request, empleado_id=None):
    """Generar certificación laboral"""
    empleado = None
    if empleado_id:
        empleado = get_object_or_404(Empleado, pk=empleado_id)
    
    if request.method == 'POST':
        form = CertificacionForm(request.POST)
        if form.is_valid():
            certificacion = form.save(commit=False)
            if not empleado:
                empleado = certificacion.empleado
            certificacion.empleado = empleado
            certificacion.generada_por = request.user
            certificacion.save()
            
            messages.success(request, f'Certificación generada para {empleado.get_nombre_completo()}.')
            return redirect('talento_humano:detalle_empleado', pk=empleado.pk)
    else:
        initial = {'empleado': empleado} if empleado else {}
        form = CertificacionForm(initial=initial)
    
    context = {
        'form': form,
        'empleado': empleado,
    }
    
    return render(request, 'talento_humano/certificaciones/form_certificacion.html', context)


# ============================================================================
# 3. SELECCIÓN Y TALENTO - PERFILES DE CARGO
# ============================================================================

class PerfilCargoListView(LoginRequiredMixin, ListView):
    """Lista de perfiles de cargo"""
    model = PerfilCargo
    template_name = 'talento_humano/seleccion/lista_perfiles_cargo.html'
    context_object_name = 'perfiles'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = PerfilCargo.objects.all()
        
        # Filtrar por área
        area = self.request.GET.get('area')
        if area:
            queryset = queryset.filter(area=area)
        
        # Filtrar por activos
        if self.request.GET.get('activos'):
            queryset = queryset.filter(activo=True)
        
        return queryset.order_by('nombre_cargo')


class PerfilCargoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo perfil de cargo"""
    model = PerfilCargo
    form_class = PerfilCargoForm
    template_name = 'talento_humano/seleccion/form_perfil_cargo.html'
    success_url = reverse_lazy('talento_humano:lista_perfiles_cargo')
    
    def form_valid(self, form):
        messages.success(self.request, 'Perfil de cargo creado exitosamente.')
        return super().form_valid(form)


class PerfilCargoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de perfil de cargo"""
    model = PerfilCargo
    template_name = 'talento_humano/seleccion/detalle_perfil_cargo.html'
    context_object_name = 'perfil'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Vacantes asociadas a este perfil
        context['vacantes'] = Vacante.objects.filter(perfil_cargo=self.object).order_by('-fecha_publicacion')
        return context


class PerfilCargoUpdateView(LoginRequiredMixin, UpdateView):
    """Editar perfil de cargo"""
    model = PerfilCargo
    form_class = PerfilCargoForm
    template_name = 'talento_humano/seleccion/form_perfil_cargo.html'
    
    def get_success_url(self):
        return reverse_lazy('talento_humano:detalle_perfil_cargo', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Perfil de cargo actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# VACANTES
# ============================================================================

class VacanteListView(LoginRequiredMixin, ListView):
    """Lista de vacantes"""
    model = Vacante
    template_name = 'talento_humano/seleccion/lista_vacantes.html'
    context_object_name = 'vacantes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Vacante.objects.all().select_related('perfil_cargo')
        
        # Filtrar por estado
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        return queryset.order_by('-fecha_publicacion')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['vacantes_abiertas'] = Vacante.objects.filter(estado='abierta').count()
        context['vacantes_cerradas'] = Vacante.objects.filter(estado='cerrada').count()
        return context


class VacanteCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva vacante"""
    model = Vacante
    form_class = VacanteForm
    template_name = 'talento_humano/seleccion/form_vacante.html'
    success_url = reverse_lazy('talento_humano:lista_vacantes')
    
    def form_valid(self, form):
        messages.success(self.request, 'Vacante creada exitosamente.')
        return super().form_valid(form)


class VacanteDetailView(LoginRequiredMixin, DetailView):
    """Detalle de vacante"""
    model = Vacante
    template_name = 'talento_humano/seleccion/detalle_vacante.html'
    context_object_name = 'vacante'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Candidatos aplicados a esta vacante
        context['procesos'] = ProcesoSeleccion.objects.filter(
            vacante=self.object
        ).select_related('candidato').order_by('-fecha_aplicacion')
        return context


class VacanteUpdateView(LoginRequiredMixin, UpdateView):
    """Editar vacante"""
    model = Vacante
    form_class = VacanteForm
    template_name = 'talento_humano/seleccion/form_vacante.html'
    
    def get_success_url(self):
        return reverse_lazy('talento_humano:detalle_vacante', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Vacante actualizada exitosamente.')
        return super().form_valid(form)


# ============================================================================
# CANDIDATOS
# ============================================================================

class CandidatoListView(LoginRequiredMixin, ListView):
    """Lista de candidatos"""
    model = Candidato
    template_name = 'talento_humano/seleccion/lista_candidatos.html'
    context_object_name = 'candidatos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Candidato.objects.all()
        
        # Búsqueda
        busqueda = self.request.GET.get('busqueda')
        if busqueda:
            queryset = queryset.filter(
                Q(nombre__icontains=busqueda) |
                Q(apellido__icontains=busqueda) |
                Q(email__icontains=busqueda) |
                Q(telefono__icontains=busqueda)
            )
        
        return queryset.order_by('-fecha_registro')


class CandidatoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo candidato"""
    model = Candidato
    form_class = CandidatoForm
    template_name = 'talento_humano/seleccion/form_candidato.html'
    success_url = reverse_lazy('talento_humano:lista_candidatos')
    
    def form_valid(self, form):
        messages.success(self.request, 'Candidato registrado exitosamente.')
        return super().form_valid(form)


class CandidatoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de candidato"""
    model = Candidato
    template_name = 'talento_humano/seleccion/detalle_candidato.html'
    context_object_name = 'candidato'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Procesos de selección del candidato
        context['procesos'] = ProcesoSeleccion.objects.filter(
            candidato=self.object
        ).select_related('vacante').order_by('-fecha_aplicacion')
        return context


# ============================================================================
# PROCESO DE SELECCIÓN
# ============================================================================

@login_required
def actualizar_etapa_candidato(request, proceso_id):
    """Actualizar etapa del proceso de selección"""
    proceso = get_object_or_404(ProcesoSeleccion, pk=proceso_id)
    
    if request.method == 'POST':
        form = ProcesoSeleccionForm(request.POST, instance=proceso)
        if form.is_valid():
            form.save()
            messages.success(request, f'Etapa actualizada para {proceso.candidato.nombre} {proceso.candidato.apellido}.')
            return redirect('talento_humano:detalle_vacante', pk=proceso.vacante.pk)
    else:
        form = ProcesoSeleccionForm(instance=proceso)
    
    context = {
        'form': form,
        'proceso': proceso,
    }
    
    return render(request, 'talento_humano/seleccion/actualizar_etapa.html', context)


# ============================================================================
# 4. CAPACITACIÓN Y DESARROLLO
# ============================================================================

class CapacitacionListView(LoginRequiredMixin, ListView):
    """Lista de capacitaciones"""
    model = Capacitacion
    template_name = 'talento_humano/capacitacion/lista_capacitaciones.html'
    context_object_name = 'capacitaciones'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Capacitacion.objects.all()
        
        # Filtrar por estado
        if self.request.GET.get('completadas'):
            queryset = queryset.filter(completada=True)
        elif self.request.GET.get('pendientes'):
            queryset = queryset.filter(completada=False, fecha_programada__gte=date.today())
        
        return queryset.order_by('-fecha_programada')


class CapacitacionCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva capacitación"""
    model = Capacitacion
    form_class = CapacitacionForm
    template_name = 'talento_humano/capacitacion/form_capacitacion.html'
    success_url = reverse_lazy('talento_humano:lista_capacitaciones')
    
    def form_valid(self, form):
        messages.success(self.request, 'Capacitación creada exitosamente.')
        return super().form_valid(form)


class CapacitacionDetailView(LoginRequiredMixin, DetailView):
    """Detalle de capacitación"""
    model = Capacitacion
    template_name = 'talento_humano/capacitacion/detalle_capacitacion.html'
    context_object_name = 'capacitacion'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Empleados inscritos
        context['inscritos'] = InscripcionCapacitacion.objects.filter(
            capacitacion=self.object
        ).select_related('empleado').order_by('fecha_inscripcion')
        return context


class CapacitacionUpdateView(LoginRequiredMixin, UpdateView):
    """Editar capacitación"""
    model = Capacitacion
    form_class = CapacitacionForm
    template_name = 'talento_humano/capacitacion/form_capacitacion.html'
    
    def get_success_url(self):
        return reverse_lazy('talento_humano:detalle_capacitacion', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Capacitación actualizada exitosamente.')
        return super().form_valid(form)


@login_required
def inscribir_capacitacion(request, capacitacion_id, empleado_id):
    """Inscribir empleado a capacitación"""
    capacitacion = get_object_or_404(Capacitacion, pk=capacitacion_id)
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    
    # Verificar si ya está inscrito
    if InscripcionCapacitacion.objects.filter(capacitacion=capacitacion, empleado=empleado).exists():
        messages.warning(request, f'{empleado.get_nombre_completo()} ya está inscrito en esta capacitación.')
    else:
        InscripcionCapacitacion.objects.create(
            capacitacion=capacitacion,
            empleado=empleado,
            fecha_inscripcion=date.today()
        )
        messages.success(request, f'{empleado.get_nombre_completo()} inscrito exitosamente.')
    
    return redirect('talento_humano:detalle_capacitacion', pk=capacitacion_id)


# ============================================================================
# 5. SEGURIDAD Y SALUD EN EL TRABAJO (SST)
# ============================================================================

@login_required
def dashboard_sst(request):
    """Dashboard de Seguridad y Salud en el Trabajo"""
    
    total_empleados = Empleado.objects.filter(estado='activo').count()
    
    # Exámenes médicos
    examenes_al_dia = ExamenMedico.objects.filter(
        fecha__gte=date.today() - timedelta(days=365),
        empleado__estado='activo'
    ).values('empleado').distinct().count()
    
    examenes_vencidos = total_empleados - examenes_al_dia
    
    # Accidentes del año
    accidentes_año = AccidenteTrabajo.objects.filter(
        fecha_accidente__year=date.today().year
    ).count()
    
    # Riesgos
    riesgos_criticos = MatrizRiesgo.objects.filter(nivel_riesgo='critico').count()
    riesgos_alto = MatrizRiesgo.objects.filter(nivel_riesgo='alto').count()
    riesgos_muy_alto = MatrizRiesgo.objects.filter(nivel_riesgo='muy alto').count()
    
    # EPP bajo stock
    epp_bajo_stock = ElementoProteccion.objects.filter(
        stock_actual__lte=F('stock_minimo'),
        activo=True
    ).count()
    
    # Próximas entregas
    proximas_entregas = EntregaEPP.objects.filter(
        fecha_entrega__gte=date.today(),
        fecha_entrega__lte=date.today() + timedelta(days=7)
    ).count()
    
    context = {
        'total_empleados': total_empleados,
        'examenes_al_dia': examenes_al_dia,
        'examenes_vencidos': examenes_vencidos,
        'accidentes_año': accidentes_año,
        'riesgos_criticos': riesgos_criticos,
        'riesgos_alto': riesgos_alto,
        'riesgos_muy_alto': riesgos_muy_alto,
        'epp_bajo_stock': epp_bajo_stock,
        'proximas_entregas': proximas_entregas,
    }
    
    return render(request, 'talento_humano/sst/dashboard_sst.html', context)


# ============================================================================
# MATRIZ DE RIESGOS
# ============================================================================

class MatrizRiesgoListView(LoginRequiredMixin, ListView):
    """Lista de riesgos identificados"""
    model = MatrizRiesgo
    template_name = 'talento_humano/sst/matriz_riesgos.html'
    context_object_name = 'riesgos'
    
    def get_queryset(self):
        queryset = MatrizRiesgo.objects.all()
        
        # Filtrar por nivel de riesgo
        nivel = self.request.GET.get('nivel')
        if nivel:
            queryset = queryset.filter(nivel_riesgo=nivel)
        
        return queryset.order_by('-nivel_riesgo', 'proceso')


class MatrizRiesgoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo riesgo"""
    model = MatrizRiesgo
    form_class = MatrizRiesgoForm
    template_name = 'talento_humano/sst/form_riesgo.html'
    success_url = reverse_lazy('talento_humano:matriz_riesgos')
    
    def form_valid(self, form):
        messages.success(self.request, 'Riesgo registrado exitosamente.')
        return super().form_valid(form)


class MatrizRiesgoUpdateView(LoginRequiredMixin, UpdateView):
    """Editar riesgo"""
    model = MatrizRiesgo
    form_class = MatrizRiesgoForm
    template_name = 'talento_humano/sst/form_riesgo.html'
    success_url = reverse_lazy('talento_humano:matriz_riesgos')
    
    def form_valid(self, form):
        messages.success(self.request, 'Riesgo actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# EXÁMENES MÉDICOS
# ============================================================================

class ExamenMedicoListView(LoginRequiredMixin, ListView):
    """Lista de exámenes médicos"""
    model = ExamenMedico
    template_name = 'talento_humano/sst/lista_examenes_medicos.html'
    context_object_name = 'examenes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = ExamenMedico.objects.all().select_related('empleado')
        
        # Filtrar por tipo
        tipo = self.request.GET.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo_examen=tipo)
        
        # Filtrar por empleado
        empleado_id = self.request.GET.get('empleado')
        if empleado_id:
            queryset = queryset.filter(empleado_id=empleado_id)
        
        return queryset.order_by('-fecha')


class ExamenMedicoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo examen médico"""
    model = ExamenMedico
    form_class = ExamenMedicoForm
    template_name = 'talento_humano/sst/form_examen_medico.html'
    success_url = reverse_lazy('talento_humano:lista_examenes_medicos')
    
    def form_valid(self, form):
        messages.success(self.request, 'Examen médico registrado exitosamente.')
        return super().form_valid(form)


class ExamenMedicoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de examen médico"""
    model = ExamenMedico
    template_name = 'talento_humano/sst/detalle_examen_medico.html'
    context_object_name = 'examen'


# ============================================================================
# ACCIDENTES DE TRABAJO
# ============================================================================

class AccidenteTrabajoListView(LoginRequiredMixin, ListView):
    """Lista de accidentes de trabajo"""
    model = AccidenteTrabajo
    template_name = 'talento_humano/sst/lista_accidentes.html'
    context_object_name = 'accidentes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = AccidenteTrabajo.objects.all().select_related('empleado')
        
        # Filtrar por gravedad
        gravedad = self.request.GET.get('gravedad')
        if gravedad:
            queryset = queryset.filter(gravedad=gravedad)
        
        return queryset.order_by('-fecha_accidente')


class AccidenteTrabajoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo accidente de trabajo"""
    model = AccidenteTrabajo
    form_class = AccidenteTrabajoForm
    template_name = 'talento_humano/sst/form_accidente.html'
    success_url = reverse_lazy('talento_humano:lista_accidentes')
    
    def form_valid(self, form):
        messages.success(self.request, 'Accidente registrado exitosamente.')
        return super().form_valid(form)


class AccidenteTrabajoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de accidente de trabajo"""
    model = AccidenteTrabajo
    template_name = 'talento_humano/sst/detalle_accidente.html'
    context_object_name = 'accidente'


class AccidenteTrabajoUpdateView(LoginRequiredMixin, UpdateView):
    """Editar accidente de trabajo"""
    model = AccidenteTrabajo
    form_class = AccidenteTrabajoForm
    template_name = 'talento_humano/sst/form_accidente.html'
    
    def get_success_url(self):
        return reverse_lazy('talento_humano:detalle_accidente', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Accidente actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# ELEMENTOS DE PROTECCIÓN PERSONAL (EPP)
# ============================================================================

class ElementoProteccionListView(LoginRequiredMixin, ListView):
    """Lista de elementos de protección personal"""
    model = ElementoProteccion
    template_name = 'talento_humano/sst/lista_epp.html'
    context_object_name = 'elementos'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['elementos_bajo_stock'] = ElementoProteccion.objects.filter(
            stock_actual__lte=F('stock_minimo'),
            activo=True
        )
        return context


class ElementoProteccionCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo elemento de protección"""
    model = ElementoProteccion
    form_class = ElementoProteccionForm
    template_name = 'talento_humano/sst/form_epp.html'
    success_url = reverse_lazy('talento_humano:lista_epp')
    
    def form_valid(self, form):
        messages.success(self.request, 'Elemento de protección creado exitosamente.')
        return super().form_valid(form)


class ElementoProteccionUpdateView(LoginRequiredMixin, UpdateView):
    """Editar elemento de protección"""
    model = ElementoProteccion
    form_class = ElementoProteccionForm
    template_name = 'talento_humano/sst/form_epp.html'
    success_url = reverse_lazy('talento_humano:lista_epp')
    
    def form_valid(self, form):
        messages.success(self.request, 'Elemento de protección actualizado exitosamente.')
        return super().form_valid(form)


class EntregaEPPListView(LoginRequiredMixin, ListView):
    """Lista de entregas de EPP"""
    model = EntregaEPP
    template_name = 'talento_humano/sst/lista_entregas_epp.html'
    context_object_name = 'entregas'
    paginate_by = 20
    
    def get_queryset(self):
        return EntregaEPP.objects.all().select_related('empleado', 'elemento').order_by('-fecha_entrega')


class EntregaEPPCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva entrega de EPP"""
    model = EntregaEPP
    form_class = EntregaEPPForm
    template_name = 'talento_humano/sst/form_entrega_epp.html'
    success_url = reverse_lazy('talento_humano:lista_entregas_epp')
    
    def form_valid(self, form):
        entrega = form.save()
        
        # Actualizar stock del elemento
        elemento = entrega.elemento
        elemento.stock_actual -= entrega.cantidad
        elemento.save()
        
        messages.success(self.request, f'Entrega registrada. Stock actualizado: {elemento.stock_actual}')
        return super().form_valid(form)


# ============================================================================
# 6. BIENESTAR LABORAL
# ============================================================================

class ActividadBienestarListView(LoginRequiredMixin, ListView):
    """Lista de actividades de bienestar"""
    model = ActividadBienestar
    template_name = 'talento_humano/bienestar/lista_actividades.html'
    context_object_name = 'actividades'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = ActividadBienestar.objects.all()
        
        tipo = self.request.GET.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        
        return queryset.order_by('-fecha_evento')


class ActividadBienestarCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva actividad de bienestar"""
    model = ActividadBienestar
    form_class = ActividadBienestarForm
    template_name = 'talento_humano/bienestar/form_actividad.html'
    success_url = reverse_lazy('talento_humano:lista_actividades_bienestar')
    
    def form_valid(self, form):
        messages.success(self.request, 'Actividad de bienestar creada exitosamente.')
        return super().form_valid(form)


class ActividadBienestarDetailView(LoginRequiredMixin, DetailView):
    """Detalle de actividad de bienestar"""
    model = ActividadBienestar
    template_name = 'talento_humano/bienestar/detalle_actividad.html'
    context_object_name = 'actividad'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['empleados_inscritos'] = self.object.empleados_inscritos.all()
        context['total_inscritos'] = self.object.get_inscritos()
        return context


@login_required
def inscribir_actividad_bienestar(request, actividad_id, empleado_id):
    """Inscribir empleado a actividad de bienestar"""
    actividad = get_object_or_404(ActividadBienestar, pk=actividad_id)
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    
    # Verificar cupos
    if actividad.cupo_maximo and actividad.get_inscritos() >= actividad.cupo_maximo:
        messages.error(request, 'No hay cupos disponibles.')
        return redirect('talento_humano:lista_actividades_bienestar')
    
    # Inscribir
    actividad.empleados_inscritos.add(empleado)
    messages.success(request, f'{empleado.get_nombre_completo()} inscrito en {actividad.nombre}.')
    
    return redirect('talento_humano:detalle_actividad_bienestar', pk=actividad_id)


# ============================================================================
# ENCUESTAS DE CLIMA ORGANIZACIONAL
# ============================================================================

class EncuestaClimaListView(LoginRequiredMixin, ListView):
    """Lista de encuestas de clima organizacional"""
    model = EncuestaClimaOrganizacional
    template_name = 'talento_humano/bienestar/lista_encuestas.html'
    context_object_name = 'encuestas'
    
    def get_queryset(self):
        return EncuestaClimaOrganizacional.objects.all().order_by('-fecha_inicio')


class EncuestaClimaCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva encuesta de clima"""
    model = EncuestaClimaOrganizacional
    form_class = EncuestaClimaForm
    template_name = 'talento_humano/bienestar/form_encuesta.html'
    success_url = reverse_lazy('talento_humano:lista_encuestas_clima')
    
    def form_valid(self, form):
        messages.success(self.request, 'Encuesta de clima creada exitosamente.')
        return super().form_valid(form)


@login_required
def responder_encuesta_clima(request, encuesta_id):
    """Responder encuesta de clima organizacional"""
    encuesta = get_object_or_404(EncuestaClimaOrganizacional, pk=encuesta_id)
    
    # Verificar si la encuesta está activa
    if not encuesta.activa or date.today() > encuesta.fecha_fin:
        messages.error(request, 'Esta encuesta ya no está disponible.')
        return redirect('talento_humano:dashboard')
    
    # Verificar si ya respondió (si no es anónima)
    if not encuesta.anonima:
        empleado = request.user.empleado
        if RespuestaEncuesta.objects.filter(encuesta=encuesta, empleado=empleado).exists():
            messages.warning(request, 'Ya has respondido esta encuesta.')
            return redirect('talento_humano:dashboard')
    
    if request.method == 'POST':
        form = RespuestaEncuestaForm(request.POST)
        if form.is_valid():
            respuesta = form.save(commit=False)
            respuesta.encuesta = encuesta
            
            if not encuesta.anonima:
                respuesta.empleado = request.user.empleado
            
            respuesta.save()
            messages.success(request, 'Gracias por responder la encuesta.')
            return redirect('talento_humano:dashboard')
    else:
        form = RespuestaEncuestaForm()
    
    context = {
        'form': form,
        'encuesta': encuesta,
    }
    
    return render(request, 'talento_humano/bienestar/responder_encuesta.html', context)


@login_required
def resultados_encuesta_clima(request, pk):
    """Ver resultados de encuesta de clima"""
    encuesta = get_object_or_404(EncuestaClimaOrganizacional, pk=pk)
    
    # Estadísticas
    respuestas = RespuestaEncuesta.objects.filter(encuesta=encuesta)
    total_respuestas = respuestas.count()
    
    # Promedios por dimensión
    promedios = {
        'ambiente_trabajo': respuestas.aggregate(Avg('puntuacion_ambiente_trabajo'))['puntuacion_ambiente_trabajo__avg'] or 0,
        'liderazgo': respuestas.aggregate(Avg('puntuacion_liderazgo'))['puntuacion_liderazgo__avg'] or 0,
        'comunicacion': respuestas.aggregate(Avg('puntuacion_comunicacion'))['puntuacion_comunicacion__avg'] or 0,
        'reconocimiento': respuestas.aggregate(Avg('puntuacion_reconocimiento'))['puntuacion_reconocimiento__avg'] or 0,
    }
    
    promedio_general = sum(promedios.values()) / len(promedios) if promedios else 0
    
    context = {
        'encuesta': encuesta,
        'total_respuestas': total_respuestas,
        'promedios': promedios,
        'promedio_general': round(promedio_general, 2),
    }
    
    return render(request, 'talento_humano/bienestar/resultados_encuesta.html', context)


# ============================================================================
# 7. GESTIÓN Y RELACIONES LABORALES - EVALUACIONES
# ============================================================================

class EvaluacionDesempeñoListView(LoginRequiredMixin, ListView):
    """Lista de evaluaciones de desempeño"""
    model = EvaluacionDesempeño
    template_name = 'talento_humano/gestion/lista_evaluaciones.html'
    context_object_name = 'evaluaciones'
    paginate_by = 20
    
    def get_queryset(self):
        return EvaluacionDesempeño.objects.all().select_related('empleado', 'evaluador').order_by('-fecha_evaluacion')


class EvaluacionDesempeñoCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva evaluación de desempeño"""
    model = EvaluacionDesempeño
    form_class = EvaluacionDesempeñoForm
    template_name = 'talento_humano/gestion/form_evaluacion.html'
    success_url = reverse_lazy('talento_humano:lista_evaluaciones')
    
    def form_valid(self, form):
        form.instance.evaluador = self.request.user
        messages.success(self.request, 'Evaluación de desempeño creada exitosamente.')
        return super().form_valid(form)


class EvaluacionDesempeñoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de evaluación de desempeño"""
    model = EvaluacionDesempeño
    template_name = 'talento_humano/gestion/detalle_evaluacion.html'
    context_object_name = 'evaluacion'


# ============================================================================
# PERMISOS
# ============================================================================

class PermisoListView(LoginRequiredMixin, ListView):
    """Lista de permisos y licencias"""
    model = Permiso
    template_name = 'talento_humano/gestion/lista_permisos.html'
    context_object_name = 'permisos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Permiso.objects.all().select_related('empleado', 'aprobado_por')
        
        # Filtrar por estado
        estado = self.request.GET.get('estado')
        if estado == 'pendientes':
            queryset = queryset.filter(aprobado=False)
        elif estado == 'aprobados':
            queryset = queryset.filter(aprobado=True)
        
        return queryset.order_by('-fecha_solicitud')


class PermisoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo permiso"""
    model = Permiso
    form_class = PermisoForm
    template_name = 'talento_humano/gestion/form_permiso.html'
    success_url = reverse_lazy('talento_humano:lista_permisos')
    
    def form_valid(self, form):
        messages.success(self.request, 'Permiso solicitado exitosamente.')
        return super().form_valid(form)


@login_required
def aprobar_permiso(request, pk):
    """Aprobar permiso"""
    permiso = get_object_or_404(Permiso, pk=pk)
    
    permiso.aprobado = True
    permiso.aprobado_por = request.user
    permiso.fecha_aprobacion = timezone.now()
    permiso.save()
    
    messages.success(request, f'Permiso aprobado para {permiso.empleado.get_nombre_completo()}.')
    return redirect('talento_humano:lista_permisos')


@login_required
def rechazar_permiso(request, pk):
    """Rechazar permiso"""
    permiso = get_object_or_404(Permiso, pk=pk)
    
    permiso.aprobado = False
    permiso.aprobado_por = request.user
    permiso.fecha_aprobacion = timezone.now()
    permiso.save()
    
    messages.warning(request, f'Permiso rechazado para {permiso.empleado.get_nombre_completo()}.')
    return redirect('talento_humano:lista_permisos')


# ============================================================================
# VACACIONES
# ============================================================================

class VacacionListView(LoginRequiredMixin, ListView):
    """Lista de vacaciones"""
    model = Vacacion
    template_name = 'talento_humano/gestion/lista_vacaciones.html'
    context_object_name = 'vacaciones'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Vacacion.objects.all().select_related('empleado', 'aprobada_por')
        
        estado = self.request.GET.get('estado')
        if estado == 'pendientes':
            queryset = queryset.filter(aprobada=False)
        elif estado == 'aprobadas':
            queryset = queryset.filter(aprobada=True)
        
        return queryset.order_by('-fecha_solicitud')


class VacacionCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva solicitud de vacaciones"""
    model = Vacacion
    form_class = VacacionForm
    template_name = 'talento_humano/gestion/form_vacacion.html'
    success_url = reverse_lazy('talento_humano:lista_vacaciones')
    
    def form_valid(self, form):
        messages.success(self.request, 'Solicitud de vacaciones creada exitosamente.')
        return super().form_valid(form)


@login_required
def aprobar_vacacion(request, pk):
    """Aprobar vacación"""
    vacacion = get_object_or_404(Vacacion, pk=pk)
    
    vacacion.aprobada = True
    vacacion.aprobada_por = request.user
    vacacion.fecha_aprobacion = timezone.now()
    vacacion.save()
    
    messages.success(request, f'Vacación aprobada para {vacacion.empleado.get_nombre_completo()}.')
    return redirect('talento_humano:lista_vacaciones')


# ============================================================================
# INCAPACIDADES
# ============================================================================

class IncapacidadListView(LoginRequiredMixin, ListView):
    """Lista de incapacidades"""
    model = Incapacidad
    template_name = 'talento_humano/gestion/lista_incapacidades.html'
    context_object_name = 'incapacidades'
    paginate_by = 20
    
    def get_queryset(self):
        return Incapacidad.objects.all().select_related('empleado').order_by('-fecha_inicio')


class IncapacidadCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva incapacidad"""
    model = Incapacidad
    form_class = IncapacidadForm
    template_name = 'talento_humano/gestion/form_incapacidad.html'
    success_url = reverse_lazy('talento_humano:lista_incapacidades')
    
    def form_valid(self, form):
        messages.success(self.request, 'Incapacidad registrada exitosamente.')
        return super().form_valid(form)


class IncapacidadDetailView(LoginRequiredMixin, DetailView):
    """Detalle de incapacidad"""
    model = Incapacidad
    template_name = 'talento_humano/gestion/detalle_incapacidad.html'
    context_object_name = 'incapacidad'


# ============================================================================
# MEMORANDOS
# ============================================================================

class MemorandoListView(LoginRequiredMixin, ListView):
    """Lista de memorandos y llamados de atención"""
    model = Memorando
    template_name = 'talento_humano/gestion/lista_memorandos.html'
    context_object_name = 'memorandos'
    paginate_by = 20
    
    def get_queryset(self):
        return Memorando.objects.all().select_related('empleado', 'emitido_por').order_by('-fecha')


class MemorandoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo memorando"""
    model = Memorando
    form_class = MemorandoForm
    template_name = 'talento_humano/gestion/form_memorando.html'
    success_url = reverse_lazy('talento_humano:lista_memorandos')
    
    def form_valid(self, form):
        form.instance.emitido_por = self.request.user
        messages.success(self.request, 'Memorando creado exitosamente.')
        return super().form_valid(form)


class MemorandoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de memorando"""
    model = Memorando
    template_name = 'talento_humano/gestion/detalle_memorando.html'
    context_object_name = 'memorando'


# ============================================================================
# REGLAMENTO INTERNO
# ============================================================================

class ReglamentoInternoListView(LoginRequiredMixin, ListView):
    """Lista de reglamentos internos"""
    model = ReglamentoInterno
    template_name = 'talento_humano/gestion/lista_reglamentos.html'
    context_object_name = 'reglamentos'
    
    def get_queryset(self):
        return ReglamentoInterno.objects.filter(activo=True).order_by('-fecha_vigencia')


class ReglamentoInternoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de reglamento interno"""
    model = ReglamentoInterno
    template_name = 'talento_humano/gestion/detalle_reglamento.html'
    context_object_name = 'reglamento'


# ============================================================================
# REPORTES Y EXPORTACIONES
# ============================================================================

@login_required
def exportar_empleados_excel(request):
    """Exportar listado de empleados a Excel"""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Encabezados
    headers = [
        'Documento', 'Nombres', 'Apellidos', 'Cargo', 'Área', 
        'Fecha Ingreso', 'Estado', 'Teléfono', 'Email'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    empleados = Empleado.objects.all().order_by('primer_apellido')
    
    for row_num, empleado in enumerate(empleados, 2):
        ws.cell(row=row_num, column=1, value=empleado.numero_documento)
        ws.cell(row=row_num, column=2, value=f"{empleado.primer_nombre} {empleado.segundo_nombre or ''}")
        ws.cell(row=row_num, column=3, value=f"{empleado.primer_apellido} {empleado.segundo_apellido or ''}")
        ws.cell(row=row_num, column=4, value=empleado.cargo)
        ws.cell(row=row_num, column=5, value=empleado.area)
        ws.cell(row=row_num, column=6, value=empleado.fecha_ingreso.strftime('%Y-%m-%d') if empleado.fecha_ingreso else '')
        ws.cell(row=row_num, column=7, value=empleado.get_estado_display())
        ws.cell(row=row_num, column=8, value=empleado.telefono)
        ws.cell(row=row_num, column=9, value=empleado.email)
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=empleados_{date.today()}.xlsx'
    
    wb.save(response)
    return response


@login_required
def reporte_nomina(request):
    """Reporte de nómina"""
    empleados_activos = Empleado.objects.filter(estado='activo').select_related('eps', 'afp', 'arl')
    
    # Calcular totales
    total_salarios = sum(emp.get_contrato_actual().salario_basico if emp.get_contrato_actual() else 0 
                         for emp in empleados_activos)
    
    context = {
        'empleados': empleados_activos,
        'total_salarios': total_salarios,
        'fecha_generacion': date.today(),
    }
    
    return render(request, 'talento_humano/reportes/reporte_nomina.html', context)


@login_required
def reporte_sst(request):
    """Reporte de Seguridad y Salud en el Trabajo"""
    
    # Estadísticas del año actual
    año_actual = date.today().year
    
    # Accidentes
    accidentes = AccidenteTrabajo.objects.filter(fecha_accidente__year=año_actual)
    total_accidentes = accidentes.count()
    accidentes_leves = accidentes.filter(gravedad='leve').count()
    accidentes_graves = accidentes.filter(gravedad='grave').count()
    accidentes_mortales = accidentes.filter(gravedad='mortal').count()
    
    # Exámenes médicos
    examenes_año = ExamenMedico.objects.filter(fecha__year=año_actual)
    
    # Riesgos
    riesgos_activos = MatrizRiesgo.objects.all()
    riesgos_criticos = riesgos_activos.filter(nivel_riesgo='critico').count()
    riesgos_altos = riesgos_activos.filter(nivel_riesgo='alto').count()
    
    # EPP
    entregas_epp = EntregaEPP.objects.filter(fecha_entrega__year=año_actual)
    
    context = {
        'año': año_actual,
        'total_accidentes': total_accidentes,
        'accidentes_leves': accidentes_leves,
        'accidentes_graves': accidentes_graves,
        'accidentes_mortales': accidentes_mortales,
        'total_examenes': examenes_año.count(),
        'riesgos_criticos': riesgos_criticos,
        'riesgos_altos': riesgos_altos,
        'total_entregas_epp': entregas_epp.count(),
        'fecha_generacion': date.today(),
    }
    
    return render(request, 'talento_humano/reportes/reporte_sst.html', context)


# ============================================================================
# API / AJAX
# ============================================================================

@login_required
def empleado_info_json(request, pk):
    """Obtener información de empleado en JSON"""
    empleado = get_object_or_404(Empleado, pk=pk)
    
    data = {
        'id': empleado.id,
        'nombre_completo': empleado.get_nombre_completo(),
        'documento': empleado.numero_documento,
        'cargo': empleado.cargo,
        'area': empleado.area,
        'email': empleado.email,
        'telefono': empleado.telefono,
        'estado': empleado.estado,
    }
    
    # Agregar info de contrato actual si existe
    contrato = empleado.get_contrato_actual()
    if contrato:
        data['contrato'] = {
            'tipo': contrato.tipo_contrato,
            'salario': str(contrato.salario_basico),
            'fecha_inicio': contrato.fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': contrato.fecha_fin.strftime('%Y-%m-%d') if contrato.fecha_fin else None,
        }
    
    return JsonResponse(data)


@login_required
def calcular_liquidacion_ajax(request):
    """Calcular liquidación de empleado (AJAX)"""
    if request.method == 'POST':
        empleado_id = request.POST.get('empleado_id')
        fecha_retiro = request.POST.get('fecha_retiro')
        
        empleado = get_object_or_404(Empleado, pk=empleado_id)
        
        # Aquí iría la lógica de cálculo de liquidación
        # Por ahora retornamos un ejemplo básico
        
        contrato = empleado.get_contrato_actual()
        if not contrato:
            return JsonResponse({'error': 'No hay contrato activo'}, status=400)
        
        # Cálculos básicos (simplificado)
        salario = contrato.salario_basico
        dias_trabajados = 30  # Ejemplo
        
        cesantias = salario * dias_trabajados / 360
        intereses_cesantias = cesantias * 0.12
        prima = salario * dias_trabajados / 360
        vacaciones = salario * dias_trabajados / 720
        
        total = cesantias + intereses_cesantias + prima + vacaciones
        
        data = {
            'cesantias': str(cesantias),
            'intereses_cesantias': str(intereses_cesantias),
            'prima': str(prima),
            'vacaciones': str(vacaciones),
            'total': str(total),
        }
        
        return JsonResponse(data)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

# Add this to your views.py file (nexusone/talento_humano/views.py)

@login_required
def calendario_capacitaciones(request):
    """Vista de calendario de capacitaciones"""
    from datetime import datetime, timedelta
    
    # Obtener mes y año de los parámetros GET, o usar el actual
    año = int(request.GET.get('año', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))
    
    # Crear fecha para el primer día del mes
    primer_dia = datetime(año, mes, 1).date()
    
    # Calcular el último día del mes
    if mes == 12:
        ultimo_dia = datetime(año + 1, 1, 1).date() - timedelta(days=1)
    else:
        ultimo_dia = datetime(año, mes + 1, 1).date() - timedelta(days=1)
    
    # Obtener capacitaciones del mes
    capacitaciones = Capacitacion.objects.filter(
        fecha_programada__gte=primer_dia,
        fecha_programada__lte=ultimo_dia
    ).order_by('fecha_programada')
    
    # Crear estructura de calendario
    # Agregar días vacíos al inicio (días de la semana anterior)
    inicio_semana = primer_dia.weekday()  # 0 = Lunes, 6 = Domingo
    
    # Crear lista de días del mes con sus capacitaciones
    dias_calendario = []
    for dia_num in range(1, ultimo_dia.day + 1):
        fecha_dia = datetime(año, mes, dia_num).date()
        caps_dia = capacitaciones.filter(fecha_programada=fecha_dia)
        dias_calendario.append({
            'fecha': fecha_dia,
            'dia': dia_num,
            'capacitaciones': caps_dia
        })
    
    # Calcular mes anterior y siguiente
    if mes == 1:
        mes_anterior = {'mes': 12, 'año': año - 1}
    else:
        mes_anterior = {'mes': mes - 1, 'año': año}
    
    if mes == 12:
        mes_siguiente = {'mes': 1, 'año': año + 1}
    else:
        mes_siguiente = {'mes': mes + 1, 'año': año}
    
    # Nombres de meses en español
    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    context = {
        'año': año,
        'mes': mes,
        'mes_nombre': meses[mes - 1],
        'dias_calendario': dias_calendario,
        'inicio_semana': inicio_semana,
        'mes_anterior': mes_anterior,
        'mes_siguiente': mes_siguiente,
        'capacitaciones_mes': capacitaciones,
        'hoy': datetime.now().date(),
    }
    
    return render(request, 'talento_humano/capacitacion/calendario_capacitaciones.html', context)