# nexusone/administrativa/inventario/views.py
# ESTRUCTURA COMPLETA DEL ARCHIVO

# ============================================
# IMPORTS
# ============================================
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.http import HttpResponse
from decimal import Decimal

# Para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Modelos y formularios
from .models import Insumo, Maquinaria, Herramienta, MovimientoKardex
from .forms import InsumoForm, MaquinariaForm, HerramientaForm, MovimientoKardexForm


# ============================================
# VISTA PRINCIPAL INVENTARIO
# ============================================
def index_inventario(request):
    """Menú principal del módulo de inventario"""
    return render(request, "administrativa/inventario/index.html")


# ============================================
# INSUMOS
# ============================================
def lista_insumo(request):
    insumos = Insumo.objects.all()
    total_inventario = sum(insumo.precio_total for insumo in insumos)
    return render(
        request,
        "administrativa/inventario/insumos/lista_insumo.html",
        {"insumos": insumos, "total_inventario": total_inventario}
    )

def nuevo_insumo(request):
    if request.method == "POST":
        form = InsumoForm(request.POST)
        if form.is_valid():
            insumo = form.save(commit=False)
            insumo.iva = Decimal(request.POST.get("iva", 0))
            insumo.descuento_proveedor = Decimal(request.POST.get("descuento_proveedor", 0))
            insumo.save()
            return redirect("administrativa:inventario:lista_insumo")
    else:
        form = InsumoForm()
    return render(request, "administrativa/inventario/insumos/form_insumo.html", {"form": form})

def editar_insumo(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == "POST":
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            insumo = form.save(commit=False)
            insumo.iva = Decimal(request.POST.get("iva", 0))
            insumo.descuento_proveedor = Decimal(request.POST.get("descuento_proveedor", 0))
            insumo.save()
            return redirect("administrativa:inventario:lista_insumo")
    else:
        form = InsumoForm(instance=insumo)
    return render(request, "administrativa/inventario/insumos/form_insumo.html", {"form": form})

def eliminar_insumo(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk)
    insumo.delete()
    return redirect("administrativa:inventario:lista_insumo")


# ============================================
# MAQUINARIA (🆕 CON request.FILES)
# ============================================
def lista_maquinaria(request):
    maquinas = Maquinaria.objects.all()
    return render(request, "administrativa/inventario/maquinaria/listar_maquinaria.html", {"maquinas": maquinas})

def nueva_maquinaria(request):
    if request.method == "POST":
        # 🆕 AGREGADO request.FILES para archivos
        form = MaquinariaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("administrativa:inventario:lista_maquinaria")
    else:
        form = MaquinariaForm()
    return render(request, "administrativa/inventario/maquinaria/nueva_maquinaria.html", {"form": form})

def editar_maquinaria(request, pk):
    maquina = get_object_or_404(Maquinaria, pk=pk)
    if request.method == "POST":
        # 🆕 AGREGADO request.FILES para archivos
        form = MaquinariaForm(request.POST, request.FILES, instance=maquina)
        if form.is_valid():
            form.save()
            return redirect("administrativa:inventario:lista_maquinaria")
    else:
        form = MaquinariaForm(instance=maquina)
    return render(request, "administrativa/inventario/maquinaria/nueva_maquinaria.html", {"form": form})

def eliminar_maquinaria(request, pk):
    maquina = get_object_or_404(Maquinaria, pk=pk)
    if request.method == "POST":
        maquina.delete()
        return redirect("administrativa:inventario:lista_maquinaria")
    return render(request, "administrativa/inventario/maquinaria/nueva_maquinaria.html", {"form": None, "delete": True})


# ============================================
# HERRAMIENTAS
# ============================================
def lista_herramientas(request):
    herramientas = Herramienta.objects.all()
    return render(request, "administrativa/inventario/herramientas/listar_herramientas.html", {"herramientas": herramientas})

def nueva_herramienta(request):
    if request.method == "POST":
        form = HerramientaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("administrativa:inventario:lista_herramientas")
    else:
        form = HerramientaForm()
    return render(request, "administrativa/inventario/herramientas/nueva_herramienta.html", {"form": form})

def editar_herramienta(request, pk):
    herramienta = get_object_or_404(Herramienta, pk=pk)
    if request.method == "POST":
        form = HerramientaForm(request.POST, instance=herramienta)
        if form.is_valid():
            form.save()
            return redirect("administrativa:inventario:lista_herramientas")
    else:
        form = HerramientaForm(instance=herramienta)
    return render(request, "administrativa/inventario/herramientas/nueva_herramienta.html", {"form": form})

def eliminar_herramienta(request, pk):
    herramienta = get_object_or_404(Herramienta, pk=pk)
    if request.method == "POST":
        herramienta.delete()
        return redirect("administrativa:inventario:lista_herramientas")
    return render(request, "administrativa/inventario/herramientas/nueva_herramienta.html", {"form": None, "delete": True})


# ============================================
# KARDEX
# ============================================
def listar_kardex(request):
    """Listar todos los movimientos del Kardex."""
    movimientos = MovimientoKardex.objects.all().order_by('-fecha')
    return render(request, "administrativa/inventario/kardex/listar_kardex.html", {"movimientos": movimientos})

def registrar_movimiento_kardex(request):
    """Registrar un nuevo movimiento en el Kardex."""
    if request.method == "POST":
        form = MovimientoKardexForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            if not movimiento.fecha:
                movimiento.fecha = timezone.now()
            movimiento.save()
            return redirect("administrativa:inventario:listar_kardex")
    else:
        form = MovimientoKardexForm(initial={"fecha": timezone.now()})
    return render(request, "administrativa/inventario/kardex/nuevo_movimiento.html", {"form": form})

def editar_movimiento(request, pk):
    """Editar un movimiento existente del Kardex."""
    movimiento = get_object_or_404(MovimientoKardex, pk=pk)
    if request.method == "POST":
        form = MovimientoKardexForm(request.POST, instance=movimiento)
        if form.is_valid():
            movimiento = form.save(commit=False)
            if not movimiento.fecha:
                movimiento.fecha = timezone.now()
            movimiento.save()
            return redirect("administrativa:inventario:listar_kardex")
    else:
        form = MovimientoKardexForm(instance=movimiento)
    return render(request, "administrativa/inventario/kardex/nuevo_movimiento.html", {"form": form})

def eliminar_movimiento(request, pk):
    movimiento = get_object_or_404(MovimientoKardex, pk=pk)
    movimiento.delete()
    return redirect("administrativa:inventario:listar_kardex")


# ============================================
# IMPORTAR DESDE EXCEL
# ============================================
def importar_excel(request):
    """Vista para importar insumos desde Excel (solo hoja Inventario_Procesado)"""
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        limpiar = request.POST.get('limpiar') == 'on'
        
        if not archivo:
            messages.error(request, '❌ Debes seleccionar un archivo')
            return redirect('administrativa:inventario:importar_excel')
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, '❌ El archivo debe ser .xlsx o .xls')
            return redirect('administrativa:inventario:importar_excel')
        
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(archivo)
            
            if 'Inventario_Procesado' in wb.sheetnames:
                ws = wb['Inventario_Procesado']
            else:
                ws = wb.active
                messages.warning(request, f'⚠️ Usando hoja: {ws.title}')
            
            if limpiar:
                count = Insumo.objects.count()
                Insumo.objects.all().delete()
                messages.warning(request, f'🗑️ Se eliminaron {count} insumos existentes')
            
            creados = 0
            actualizados = 0
            errores = []
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 3:
                    continue
                    
                codigo = str(row[0]).strip() if row[0] else None
                nombre = str(row[1]).strip() if row[1] else None
                
                try:
                    stock_inicial = int(row[2]) if row[2] else 0
                except (ValueError, TypeError):
                    stock_inicial = 0
                
                if not codigo or not nombre:
                    errores.append(f'Fila {idx}: código o nombre vacío')
                    continue
                
                try:
                    insumo, created = Insumo.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': 'Importado desde Excel',
                            'unidad': 'UND',
                            'precio_unitario': Decimal('0.00'),
                            'stock_minimo': 0,
                            'stock_maximo': 1000,
                            'iva': Decimal('19.00'),
                            'descuento_proveedor': Decimal('0.00'),
                        }
                    )
                    
                    if created and stock_inicial > 0:
                        MovimientoKardex.objects.create(
                            insumo=insumo,
                            tipo='entrada',
                            cantidad=stock_inicial,
                            observacion='Stock inicial importado desde Excel',
                            fecha=timezone.now()
                        )
                        creados += 1
                    else:
                        actualizados += 1
                
                except Exception as e:
                    errores.append(f'Fila {idx} ({codigo}): {str(e)}')
            
            if creados > 0:
                messages.success(request, f'✅ {creados} insumos creados con stock inicial')
            
            if actualizados > 0:
                messages.info(request, f'🔄 {actualizados} insumos ya existían (no se modificó su stock)')
            
            if errores:
                for error in errores[:5]:
                    messages.warning(request, f'⚠️ {error}')
                if len(errores) > 5:
                    messages.warning(request, f'⚠️ ... y {len(errores) - 5} errores más')
            
            if creados > 0 or actualizados > 0:
                messages.success(request, '🎉 ¡Importación completada!')
                return redirect('administrativa:inventario:lista_insumo')
        
        except ImportError:
            messages.error(request, '❌ Instala openpyxl: pip install openpyxl')
        except Exception as e:
            messages.error(request, f'❌ Error al procesar: {str(e)}')
        
        return redirect('administrativa:inventario:importar_excel')
    
    return render(request, 'administrativa/inventario/insumos/importar_excel.html')


# ============================================
# EXPORTAR A EXCEL
# ============================================
def exportar_excel(request):
    """Exporta el inventario actual a Excel en tiempo real"""
    
    insumos = Insumo.objects.all().order_by('codigo')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos
    header_fill = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid")
    header_font = Font(bold=True, size=12, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Código', 'Nombre', 'Stock']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row_num, insumo in enumerate(insumos, 2):
        ws.cell(row=row_num, column=1, value=insumo.codigo).border = thin_border
        ws.cell(row=row_num, column=2, value=insumo.nombre).border = thin_border
        ws.cell(row=row_num, column=3, value=insumo.stock_actual).border = thin_border
    
    # Ajustar columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    
    # Preparar descarga
    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nombre_archivo = f"Inventario_{fecha_actual}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    
    return response